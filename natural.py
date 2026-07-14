#!/usr/bin/env python3
"""天然钻石部门(真诚) 工厂单 → 聚水潭 (集成版 v3)
集成: 黛宝 / 布心 / 猛哥
- 自动识别工厂
- 客户名匹配自适应新老 GIA sheet 格式 (C=商品名称 或 C=销售+D=客户)
- 单名子串匹配 (猛哥场景)

用法:
    export FS_APP_ID='cli_aa99176850795bda'
    export FS_APP_SECRET='xxx'
    python3 natural.py <工厂单.xls> --au 900 --pt 380
"""
import os, sys, re, argparse, subprocess, math
from datetime import datetime
import openpyxl
from openpyxl.styles import PatternFill
import requests

GIA_TOKEN = 'EkKhsOyrDhwHlDtH2vqcI0VQn1c'
FEISHU_BASE = 'https://open.feishu.cn/open-apis'
NUM_ORDER = re.compile(r'^\d+-+\d+-+\d+$')  # 培育钻多杠单号 (如 6-5-2)


# ==================== 通用工具 ====================
def _num(v):
    if v is None or v == '': return 0
    try: return float(str(v).replace(',', ''))
    except (ValueError, TypeError): return 0


def _cell(row, idx):
    if idx is None or idx < 0 or idx >= len(row): return None
    v = row[idx]
    return v if v not in (None, '') else None


def _norm_code(cert):
    """商品编码归一化: 双杠合并, 剥尾巴的 - / —, 加 -1 后缀."""
    if not cert: return None
    s = str(cert).strip()
    s = re.sub(r'-+', '-', s)
    s = s.rstrip('-').rstrip('—').strip()
    if not s: return None
    if re.search(r'-\d+$', s):
        return s
    return f"{s}-1"


def _current_year_gia_sheets(sheets, year=None, month=None):
    """只保留 sheet 名结尾是"目标年份"的.
       规则:
       - 默认只跑当年 (title 后缀 '26' → 2026)
       - 但当前月份 <= 3 (Q1 过渡期), 也把去年一并跑 (客户单可能是去年配的石头)
         例: 2027 年 3 月 → 跑 title 后缀 '27' 和 '26' 的所有 sheet
       - 兜底: 如果一个都没匹配上, 返回全部 (防止年份识别失效)
    """
    now = datetime.now()
    year = year if year is not None else now.year
    month = month if month is not None else now.month
    include = {str(year)[-2:]}
    if month <= 3:
        include.add(str(year - 1)[-2:])
    hits = [s for s in sheets
            if any(s.get('title', '').rstrip().endswith(y) for y in include)]
    return hits if hits else sheets


def _sort_gia_sheets(sheets, months=2):
    """GIA 订货 sheet 按 (年, 月) 元组倒排, 取最近 N 个 (0=全部).
       sheet 名格式:
         '4月订货26'  → 年=2026, 月=4
         '12月订货25' → 年=2025, 月=12
         '6月订货'    → 年=None (无年份, 排最后)
       修复: 旧版数字拼接 "1225" > "426" 导致 12月25 被误认为最新.
    """
    if months <= 0:
        return sheets

    def key(sh_with_idx):
        idx, sh = sh_with_idx
        title = sh.get('title', '')
        # M月订货YY 或 M月订货YYYY
        m = re.match(r'^\s*(\d{1,2})月订货\s*(\d{0,4})\s*$', title)
        if m:
            month = int(m.group(1))
            year_s = m.group(2)
            if year_s:
                year = int(year_s)
                # 两位年份补 2000 (25→2025, 5→2005)
                if year < 100:
                    year = 2000 + year
                return (1, year, month)  # 有年月, 排最前
            else:
                # 只有月份 (如 "6月订货") = 更老的历史, 排最后
                return (0, 0, month)
        # 完全解析不了的 → 按 sheet 顺序 (idx 越小越靠前)
        return (-1, 0, -idx)

    indexed = list(enumerate(sheets))
    sorted_ = sorted(indexed, key=key, reverse=True)
    return [sh for _, sh in sorted_[:months]]


def ensure_xlsx(path):
    if path.lower().endswith('.xlsx'):
        return path
    if not path.lower().endswith('.xls'):
        raise RuntimeError(f"扩展名不是 .xls 或 .xlsx: {path}")
    out = path.rsplit('.', 1)[0] + '_conv.xlsx'
    if os.path.exists(out):
        return out
    try:
        subprocess.run(['soffice', '--headless', '--convert-to', 'xlsx',
                        path, '--outdir', os.path.dirname(out) or '.'],
                       check=True, capture_output=True)
        return out
    except (FileNotFoundError, subprocess.CalledProcessError):
        raise RuntimeError("请用 Excel/WPS 手动打开 .xls 后另存为 .xlsx 再跑")


# ==================== 飞书客户端 ====================
class FeishuSheetClient:
    """v20.2: token 自动过期刷新
       飞书 tenant_access_token 2 小时有效期. 旧代码 self._token 永远缓存不刷新,
       导致进程活过 2 小时后 GIA 查询全部 401 失败 (客户端表现: 属性/成本全空).
    """
    def __init__(self, app_id, app_secret):
        import time as _time_mod
        self.app_id, self.app_secret = app_id, app_secret
        self._token = None
        self._token_expire = 0.0  # unix 时间戳, token 失效时间
        self._time_mod = _time_mod

    def _get_token(self):
        now = self._time_mod.time()
        # 未过期 (提前 5 分钟视作过期) → 复用
        if self._token and self._token_expire > now:
            return self._token
        r = requests.post(
            f'{FEISHU_BASE}/auth/v3/tenant_access_token/internal',
            json={'app_id': self.app_id, 'app_secret': self.app_secret},
            timeout=10)
        d = r.json()
        if d.get('code') != 0:
            raise RuntimeError(f"飞书 token 失败: {d}")
        self._token = d['tenant_access_token']
        expire_in = d.get('expire', 7200)   # 秒
        self._token_expire = now + expire_in - 300  # 提前 5 分钟刷新
        return self._token

    def _headers(self):
        return {'Authorization': f'Bearer {self._get_token()}'}

    def list_sheets(self, token):
        r = requests.get(
            f'{FEISHU_BASE}/sheets/v3/spreadsheets/{token}/sheets/query',
            headers=self._headers(), timeout=10)
        return r.json().get('data', {}).get('sheets', [])

    def read_range(self, token, range_str):
        r = requests.get(
            f'{FEISHU_BASE}/sheets/v2/spreadsheets/{token}/values/{range_str}',
            headers=self._headers(), timeout=15)
        return r.json().get('data', {}).get('valueRange', {}).get('values', []) or []


# ==================== GIA 库存查询 (自适应 layout) ====================
_GIA_CACHE = {}
_LAYOUT_CACHE = {}


def _parse_layout(header):
    """扫描表头, 识别字段位置. 支持新老两种格式:
       新 sheet: C=商品名称 D=证书 F=形状 G=主石重量 H=颜色等级 P=成本 O=商品编码
       老 sheet: C=销售 D=客户 E=证书 F=形状 G=克拉 H=颜色 O=证书编号 S=人民币
    """
    layout = {}
    for i, v in enumerate(header):
        if v is None or v == '': continue
        s = str(v).strip()
        if s == '商品名称': layout['商品名称'] = i
        elif s in ('销售', '业务', '业务员', '销售员'): layout['销售'] = i
        elif s in ('客户', '客户名', '客户名称', '姓名'): layout['客户'] = i
        elif s == '证书': layout['证书'] = i
        elif s in ('证书编号', '证书号', 'GIA号', 'GIA编号'): layout['证书编号'] = i
        elif s == '商品编码': layout['商品编码'] = i
        elif s == '成本': layout['成本'] = i
        elif s in ('人民币', '含税人民币约', '人民币约'): layout['人民币'] = i
        elif s == '形状': layout['形状'] = i
        elif s in ('主石重量', '克拉', '重量', '重量ct', '重量(ct)'): layout['克拉'] = i
        elif s in ('颜色等级', '颜色'): layout['颜色'] = i
        elif s == '净度': layout['净度'] = i
        elif s == '切工': layout['切工'] = i
        elif s == '抛光': layout['抛光'] = i
        elif s in ('对称性', '对称'): layout['对称性'] = i
        elif s == '荧光': layout['荧光'] = i
    return layout


def _load_gia_sheet(client, sheet_id):
    """v19.10: 读取范围扩大
       列: A→T (20列) 改为 A→AF (32列), 覆盖客户名/证书号/成本可能出现的所有列
       行: 500 改为 2000, 一个月订货 sheet 通常几百条, 2000 足够
    """
    if sheet_id in _GIA_CACHE:
        return _GIA_CACHE[sheet_id], _LAYOUT_CACHE[sheet_id]
    header_rows = client.read_range(GIA_TOKEN, f'{sheet_id}!A1:AF1')
    layout = _parse_layout(header_rows[0] if header_rows else [])
    rows = client.read_range(GIA_TOKEN, f'{sheet_id}!A2:AF2000')
    _GIA_CACHE[sheet_id] = rows
    _LAYOUT_CACHE[sheet_id] = layout
    return rows, layout


def _reverse_name(kuanhao):
    if '-' not in kuanhao: return None
    parts = kuanhao.split('-', 1)
    return parts[1].strip() + '-' + parts[0].strip()


def _extract_attrs(row, layout):
    def g(k):
        idx = layout.get(k)
        return _cell(row, idx) if idx is not None else None
    return {
        '证书':     str(g('证书') or '').strip(),
        '形状':     g('形状'),
        '主石重量':  g('克拉'),
        '颜色等级':  g('颜色'),
        '净度':     g('净度'),
        '切工':     g('切工'),
        '抛光':     g('抛光'),
        '对称性':    g('对称性'),
        '荧光':     g('荧光'),
    }


def _row_get_cost(row, layout):
    for k in ('成本', '人民币'):
        idx = layout.get(k)
        if idx is not None:
            return _num(_cell(row, idx))
    return 0


def _row_get_cert_no(row, layout):
    for k in ('商品编码', '证书编号'):
        idx = layout.get(k)
        if idx is not None:
            v = _cell(row, idx)
            if v is not None:
                return str(v).strip()
    return ''


def _row_match_name(row, layout, keys):
    """判断这行是否匹配 keys.
       keys 每个元素可能是:
         - 完整"客户-销售" 或 "销售-客户"
         - 单个名字 (如"丁平平") → 走子串匹配
       返回 (是否匹配, 规范化的 客户-销售 字符串)
    """
    # 收集这行的可能名字 (拆开 C 商品名称, 或者 C销售+D客户)
    row_names = set()
    combo = None

    if '商品名称' in layout:
        c_val = str(_cell(row, layout['商品名称']) or '').strip()
        if c_val:
            row_names.add(c_val)
            for part in c_val.split('-'):
                p = part.strip()
                if p: row_names.add(p)
            combo = c_val

    if '销售' in layout and '客户' in layout:
        sale = str(_cell(row, layout['销售']) or '').strip()
        cust = str(_cell(row, layout['客户']) or '').strip()
        if sale: row_names.add(sale)
        if cust: row_names.add(cust)
        if sale and cust:
            row_names.add(f'{cust}-{sale}')
            row_names.add(f'{sale}-{cust}')
            combo = f'{cust}-{sale}'

    if not row_names:
        return False, None

    for k in keys:
        if not k: continue
        # 完整精确匹配
        if k in row_names:
            return True, combo or k
        # 单名: 判断是不是这行任意名字段的完全匹配 (长度≥2 防误匹)
        if '-' not in k and len(k) >= 2:
            for name in row_names:
                # name 里可能是"客户-销售"合并串, 或者单名. 拆开每段判断
                for part in name.split('-'):
                    if part.strip() == k:
                        return True, combo or name
    return False, None


def search_gia(client, order_sheets, kuanhao, cert_no):
    """
    v20.3 重写: 用证书号锁定唯一主石, 避免多颗主石累加成 cost1.

    Step 1 (证书号锁定): 工厂单有证书号 → GIA 库存里找证书号 == 工厂号 的**唯一** 那条 → cost1
    Step 2 (客户名扫描):
        - 散货行 → cost2 累加 (cost2_count++)
        - 主石行 (证书号 != Step 1 那条) → 计数 (main_stone_extra_count++), 不累加 cost1
        - 如果 Step 1 没锁定 (无证书号 或 库存没这颗) → 用客户名第一条主石作 cost1

    警示 (红底):
        - 散货 >= 2 (客户名下多组散货)
        - 主石额外行 >= 1 (客户名下除本颗还有其他主石)
    """
    hits = []
    cost1 = 0
    cost2 = 0
    cost2_count = 0
    main_stone_extra_count = 0   # 客户名匹配到的额外主石行数 (排除 Step 1 那条)
    attrs = {}
    real_c_name = None
    main_hit_cert = None         # Step 1 锁定的证书号 (避免 Step 2 重复计算)

    keys = []
    if kuanhao:
        keys.append(kuanhao.strip())
        rev = _reverse_name(kuanhao)
        if rev: keys.append(rev)

    cert_no_str = str(cert_no or '').strip()
    cert_variants = []
    if cert_no_str:
        cert_variants.append(cert_no_str)
        # 工厂号带 LG 前缀 → 也试去掉 LG (库存可能存无 LG 版)
        if cert_no_str.upper().startswith('LG'):
            cert_variants.append(cert_no_str[2:].lstrip('_-').strip())

    # ---------- Step 1: 证书号锁定唯一主石行 ----------
    if cert_variants:
        for sh in order_sheets:
            if main_hit_cert: break
            title = sh.get('title', '')
            rows, layout = _load_gia_sheet(client, sh.get('sheet_id'))
            for row in rows:
                sheet_cert = _row_get_cert_no(row, layout)
                if sheet_cert not in cert_variants: continue
                attrs = _extract_attrs(row, layout)
                p_val = _row_get_cost(row, layout)
                d_val = attrs.get('证书', '') or ''
                cost1 = p_val
                main_hit_cert = sheet_cert
                if '商品名称' in layout:
                    real_c_name = str(_cell(row, layout['商品名称']) or '').strip()
                elif '销售' in layout and '客户' in layout:
                    sale = str(_cell(row, layout['销售']) or '').strip()
                    cust = str(_cell(row, layout['客户']) or '').strip()
                    real_c_name = f'{cust}-{sale}'
                hits.append(f'{title}(证书号锁定):{d_val}¥{p_val:.0f}(C={real_c_name})')
                break

    # ---------- Step 2: 客户名扫描 (散货累加 + 其他主石计数) ----------
    scan_keys = list(keys)
    if real_c_name and real_c_name not in scan_keys:
        scan_keys.append(real_c_name)

    if scan_keys:
        for sh in order_sheets:
            title = sh.get('title', '')
            rows, layout = _load_gia_sheet(client, sh.get('sheet_id'))
            for row in rows:
                matched, c_norm = _row_match_name(row, layout, scan_keys)
                if not matched: continue
                sheet_cert = _row_get_cert_no(row, layout)
                # 跳过 Step 1 已算过的证书号
                if main_hit_cert and sheet_cert == main_hit_cert:
                    continue

                d_val = str(_cell(row, layout.get('证书', -1)) or '').strip() if '证书' in layout else ''
                p_val = _row_get_cost(row, layout)
                if '散货' in d_val:
                    cost2 += p_val
                    cost2_count += 1
                    hits.append(f'{title}:散货¥{p_val:.0f}')
                else:
                    # 客户名的主石行
                    if cost1 == 0:
                        # Step 1 没锁定 (无证书号 或 库存里没这颗) → 用这条作 cost1
                        cost1 = p_val
                        attrs = _extract_attrs(row, layout)
                        main_hit_cert = sheet_cert or f'__cust_{c_norm}__'
                        real_c_name = real_c_name or c_norm
                        hits.append(f'{title}(客户名主石):{d_val}¥{p_val:.0f}(C={c_norm})')
                    else:
                        # 已有 cost1, 这是"额外"主石 (客户不止一颗)
                        main_stone_extra_count += 1
                        hits.append(
                            f'{title}(客户名额外主石):{d_val}¥{p_val:.0f}(证书={sheet_cert or "-"})'
                        )

    return {
        'cost1': cost1, 'cost2': cost2,
        '散货行数': cost2_count,
        '主石额外行数': main_stone_extra_count,   # v20.3
        'attrs': attrs,
        'debug': ';'.join(hits) if hits else '❌ 无匹配',
    }


# ==================== 工厂识别 ====================
def detect_factory(wb, excel_path=None):
    # 二厂真诚: 文件名含 "真诚出货单"
    if excel_path:
        fname = os.path.basename(excel_path)
        if '真诚出货单' in fname:
            return '二厂'
    names = wb.sheetnames
    if any('PT出货单' in n for n in names) and any('18K出货单' in n for n in names):
        return '布心'
    if any(n.startswith('出货单-') for n in names):
        return '黛宝'
    # 猛哥: 最后 sheet row 4 有"编码"列
    ws = wb[names[-1]]
    b4 = str(ws.cell(row=4, column=2).value or '').strip()
    c4 = str(ws.cell(row=4, column=3).value or '').strip()
    if b4 == '单号' and c4 == '编码':
        return '猛哥'
    return None


# ==================== 黛宝 解析 (天然钻, sheet 名 "出货单-*") ====================
def parse_daibao(xlsx_path, au_price, pt_price):
    """黛宝天然钻工厂单 (sheet '出货单-SG*'):
       表头 r4/r5 (真表头 r4=分类, r5=细分):
         A序号 B款号 C条码号 E证书号 G手寸 H名称 J件数 K总重
         M(13)金料1: N净重 O损耗 Q连耗金 R(18)折足金重 T金价 U金料总额
         V(22)金料2: (双段金料, 通常一段有值)
         AD(30)主石: AE(31)石重 AI(35)主钻
         AJ-CC 副石(1-10) 每 5 列 一组: 品名/数量/石重/单价/金额
           石重列: AL(38) AQ(43) AV(48) BA(53) BF(58) BK(63) BP(68) BU(73) BZ(78) CE(83)
         CJ(88)副石镶工费 CK(89)主石镶工费 CL(90)其它工艺费 CM(91)加工费 CO(93)总价

       镶嵌成本 = R折足金 × 金价 + CO总价
    """
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    target = next((n for n in wb.sheetnames if n.startswith('出货单-')), wb.sheetnames[0])
    ws = wb[target]
    # 副石 10 组的石重列位 (每组 5 列步长, 从 AL=38 起, 到 CE=83)
    SIDE_W_COLS = [38, 43, 48, 53, 58, 63, 68, 73, 78, 83]
    items = []
    for r in range(6, ws.max_row + 1):
        a = ws.cell(row=r, column=1).value
        if not isinstance(a, (int, float)): break
        kuanhao = ws.cell(row=r, column=2).value
        cert    = ws.cell(row=r, column=5).value
        ring    = ws.cell(row=r, column=7).value
        name    = ws.cell(row=r, column=8).value
        qty     = ws.cell(row=r, column=10).value or 1
        weight  = ws.cell(row=r, column=11).value
        material = str(ws.cell(row=r, column=13).value or '').strip()
        zhezu   = _num(ws.cell(row=r, column=18).value)   # R 折足金重
        co_val  = _num(ws.cell(row=r, column=93).value)   # CO 总价
        main_w  = _num(ws.cell(row=r, column=31).value)   # AE 主石石重
        side_w  = sum(_num(ws.cell(row=r, column=c).value) for c in SIDE_W_COLS)
        gp = pt_price if 'PT' in material.upper() else au_price
        cost_mount = math.ceil(float(gp) * zhezu + co_val)
        mat_disp = 'PT950' if 'PT' in material.upper() else (material.split()[0] if material else '')
        items.append({
            'no': int(a), 'row': r,
            '款号': str(kuanhao or '').strip(),
            '证书号': str(cert or '').strip(),
            '条码号': str(ws.cell(row=r, column=3).value or '').strip(),
            '品名': str(name or '').strip(),
            '手寸': str(ring or '').strip(),
            '材质颜色': mat_disp,
            '件数': int(qty) if isinstance(qty, (int, float)) else 1,
            '总重': weight,
            '主石重量_ct': main_w if main_w > 0 else None,
            '副石重量_合计': round(side_w, 4) if side_w > 0 else None,
            '镶嵌成本': cost_mount,
        })
    return items


# ==================== 布心 解析 ====================
def _find_last_batch_range(ws):
    last_合计 = None
    for r in range(ws.max_row, 0, -1):
        v = ws.cell(row=r, column=2).value
        if v and '合计' in str(v):
            last_合计 = r; break
    if not last_合计: return None
    for r in range(last_合计 - 1, 0, -1):
        v = ws.cell(row=r, column=2).value
        if v and str(v).strip() == '客户':
            return (r + 2, last_合计 - 1)
    return None


def parse_buxin(xlsx_path, au_price, pt_price):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    items = []
    no = 0
    sheet_configs = [
        ('PT出货单', 0.955, pt_price, 'PT950'),
        ('18K出货单', 0.755, au_price, '18K'),
    ]
    for key, ratio, gp, mat_disp in sheet_configs:
        sname = next((n for n in wb.sheetnames if key in n), None)
        if not sname: continue
        ws = wb[sname]
        rng = _find_last_batch_range(ws)
        if not rng: continue
        start, end = rng
        print(f"  [{sname}] 最后一批 row {start}-{end}")
        for r in range(start, end + 1):
            kuanhao = ws.cell(row=r, column=2).value
            cert    = ws.cell(row=r, column=4).value
            ring    = ws.cell(row=r, column=5).value
            name    = ws.cell(row=r, column=6).value
            qty     = ws.cell(row=r, column=7).value
            weight  = ws.cell(row=r, column=9).value
            l_val   = _num(ws.cell(row=r, column=12).value)
            am_val  = _num(ws.cell(row=r, column=39).value)
            # v19.3/v19.6: 副石列位 (布心)
            #   副石1 组: U(21)数量  V(22)石重  W(23)单价  X(24)金额 ← 天然散货 (3200元/ct)
            #   副石2 组: Y(25)数量  Z(26)石重  AA(27)单价 AB(28)金额 ← 培育散货 (1120元/ct)
            side1_w   = _num(ws.cell(row=r, column=22).value)  # V 副石1石重
            side1_amt = _num(ws.cell(row=r, column=24).value)  # X 副石1金额
            side2_w   = _num(ws.cell(row=r, column=26).value)  # Z 副石2石重
            side2_amt = _num(ws.cell(row=r, column=28).value)  # AB 副石2金额
            side_w_total = round(side1_w + side2_w, 4) if (side1_w + side2_w) > 0 else None
            if l_val <= 0: continue  # 跳过退货
            if not str(kuanhao or '').strip(): continue
            cost_mount = math.ceil(l_val * ratio * gp + am_val)
            no += 1
            items.append({
                'no': no, 'row': r,
                '款号': str(kuanhao).strip(),
                '证书号': str(cert or '').strip(),
                '条码号': '',
                '品名': str(name or '').strip(),
                '手寸': str(ring or '').strip(),
                '材质颜色': mat_disp,
                '件数': int(qty) if isinstance(qty, (int, float)) else 1,
                '总重': weight,
                '副石重量_合计': side_w_total,
                '镶嵌成本': cost_mount,
                '_副石1金额': side1_amt,
                '_副石2金额': side2_amt,
                '_sheet': sname,
            })
    return items


# ==================== 猛哥 解析 (只处理真诚部门客订) ====================
def parse_menggou(xlsx_path, au_price, pt_price):
    """猛哥工厂单 (真诚 + 培育部门混合, 结料结算):
       只处理: A 序号 = "19楼" (真诚部门) + B 单号 非空 (客户名 = 客订)
       跳过: A 非"19楼" (培育部门, 老 process.py 处理) 或 B 空 (现货)

       表头 row 4 (跟培育钻猛哥同表, 相同列位):
         A序号 B单号 C编码 D品名 E成色 F手寸 G件数 H总重量
         K(11)含耗金重 L(12)折足金 M(13)金价 N(14)金值
         P(16)配件费 R(18)主石重ct U(21)主石金额 W(23)副石重ct AA(27)副石金额
         AB(28)镶石费 AC(29)镶主石费 AD(30)版蜡 AE(31)工艺费 AF(32)工费
         AG(33)总金额

       结料算法 (跟培育钻 parse_E 一致, 猛哥是结料的):
         折足金 = L 有值? L : 含耗金重 × 比例  (PT 0.952 / K金 0.755)
         金价   = M 有值? M : (PT? pt_price : au_price)  ← 需要当天金价!
         金值   = N 有值? N : 折足金 × 金价
         镶嵌成本 = 金值 + 配件费 + 主石金额 + 副石金额 + 镶石费 + 镶主石费 + 版蜡 + 工艺费 + 工费

       AG 总金额 是工厂内部结算 (不含金子成本), 我们必须重算带金价的成本。
    """
    if not (au_price and pt_price):
        raise ValueError('猛哥是结料, 必须提供当天金价 (au_price 与 pt_price)')

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    # 最后一个 sheet 是当天出货 (跟培育钻猛哥一致)
    ws = wb[wb.sheetnames[-1]]
    items = []
    no = 0
    for r in range(6, ws.max_row + 1):
        a = ws.cell(row=r, column=1).value       # A 序号
        b = ws.cell(row=r, column=2).value       # B 单号
        d = ws.cell(row=r, column=4).value       # D 品名

        a_str = str(a or '').strip()
        b_str = str(b or '').strip()

        # 只处理真诚部门客订
        if a_str != '19楼': continue    # 不是真诚部门 → 培育部门老流程管
        if not b_str: continue          # 真诚现货 → 跳过

        material = str(ws.cell(row=r, column=5).value or '').strip()  # E 成色
        ring = ws.cell(row=r, column=6).value           # F 手寸
        qty  = ws.cell(row=r, column=7).value or 1
        weight = ws.cell(row=r, column=8).value          # H 总重量
        c_val = ws.cell(row=r, column=3).value           # C 编码

        han_hao      = _num(ws.cell(row=r, column=11).value)   # K 含耗金重
        factory_L    = ws.cell(row=r, column=12).value          # L 折足金
        factory_M    = ws.cell(row=r, column=13).value          # M 金价
        factory_N    = ws.cell(row=r, column=14).value          # N 金值
        peijian_fei  = _num(ws.cell(row=r, column=16).value)   # P 配件费
        zhushi_jin   = _num(ws.cell(row=r, column=21).value)   # U 主石金额
        fushi_jin    = _num(ws.cell(row=r, column=27).value)   # AA 副石金额
        xiangshi_fei = _num(ws.cell(row=r, column=28).value)   # AB 镶石费
        xzhushi_fei  = _num(ws.cell(row=r, column=29).value)   # AC 镶主石费
        banla        = _num(ws.cell(row=r, column=30).value)   # AD 版蜡
        gongyi       = _num(ws.cell(row=r, column=31).value)   # AE 工艺费
        gongfei      = _num(ws.cell(row=r, column=32).value)   # AF 工费
        main_w_ct    = _num(ws.cell(row=r, column=18).value)   # R 主石重ct
        side_w_ct    = _num(ws.cell(row=r, column=23).value)   # W 副石重ct

        mat_s = material.lower()
        if 'pt' in mat_s:
            ratio = 0.952
        else:
            ratio = 0.755

        # 折足金
        if isinstance(factory_L, (int, float)) and factory_L != 0:
            zhe_zu = factory_L
        elif han_hao:
            zhe_zu = round(han_hao * ratio, 5)
        else:
            zhe_zu = 0

        # 金价 (当天金价, 工厂通常不填 M/N)
        if isinstance(factory_M, (int, float)) and factory_M > 0:
            gp = factory_M
        elif 'pt' in mat_s:
            gp = pt_price
        else:
            gp = au_price

        # 金值
        if isinstance(factory_N, (int, float)) and factory_N > 0:
            jin_zhi = factory_N
        elif zhe_zu and gp:
            jin_zhi = round(zhe_zu * gp, 5)
        else:
            jin_zhi = 0

        # 镶嵌成本 = 金值 + 各项工费/石费
        cost = math.ceil(jin_zhi + peijian_fei + zhushi_jin + fushi_jin +
                          xiangshi_fei + xzhushi_fei + banla + gongyi + gongfei)

        mat_up = material.upper()
        if 'PT' in mat_up: mat_disp = 'PT950'
        elif 'K白' in mat_up: mat_disp = '18K白'
        elif 'K红' in mat_up: mat_disp = '18K红'
        elif 'K黄' in mat_up: mat_disp = '18K黄'
        elif 'K玫' in mat_up: mat_disp = '18K玫'
        else: mat_disp = material or '18K'

        no += 1
        items.append({
            'no': no, 'row': r,
            '款号': b_str,                          # 客户名
            '证书号': str(c_val or '').strip(),      # C 编码
            '条码号': '',
            '品名': str(d or '').strip(),
            '手寸': str(ring or '').strip(),
            '材质颜色': mat_disp,
            '件数': int(qty) if isinstance(qty, (int, float)) else 1,
            '总重': weight,
            '主石重量_ct': main_w_ct if main_w_ct > 0 else None,
            '副石重量_合计': side_w_ct if side_w_ct > 0 else None,
            '镶嵌成本': cost,
            '_折足金': zhe_zu,
            '_金价': gp,
            '_金值': jin_zhi,
            '_sheet': ws.title,
        })
    return items


# ==================== 二厂 真诚出货单 解析 (结价) ====================
def parse_erchang(xlsx_path, au_price, pt_price):
    """二厂 真诚出货单 (结价, AB 合计 = 镶嵌成本).
       只读最后一个 sheet (最新一天); C 单号非空=客订; 空=现货跳过.
    """
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb[wb.sheetnames[-1]]
    date_str = ''
    r3r = ws.cell(row=3, column=18).value
    if r3r and '汇总日期' in str(r3r):
        date_str = re.sub(r'^\D+', '', str(r3r)).strip()

    items = []
    no = 0
    for r in range(9, ws.max_row + 1):
        a = ws.cell(row=r, column=1).value
        if not isinstance(a, (int, float)):
            continue
        c = ws.cell(row=r, column=3).value
        c_str = str(c or '').strip()
        if not c_str:
            continue

        b = ws.cell(row=r, column=2).value
        d = ws.cell(row=r, column=4).value
        e = str(ws.cell(row=r, column=5).value or '').strip()
        f = ws.cell(row=r, column=6).value or 1
        g_wt = ws.cell(row=r, column=7).value
        # v19.8: 主石/副石重量抓取
        #   主石组 (L-O): L 石号 M 粒数 N 主石ct O 镶费
        #   辅石组 (P-T): P 石号 Q 粒数 R 辅石ct S 金额 T 微镶费   ← 相当于副石1
        #   副石2 (U-X): U 石号 V 粒数 W 副石2ct X 微镶费
        main_w  = _num(ws.cell(row=r, column=14).value)   # N 主石重ct
        side1_w = _num(ws.cell(row=r, column=18).value)   # R 辅石重ct
        side2_w = _num(ws.cell(row=r, column=23).value)   # W 副石2重ct
        side_w_total = round(side1_w + side2_w, 4) if (side1_w + side2_w) > 0 else None
        ab_total = _num(ws.cell(row=r, column=28).value)  # AB 合计 = 镶嵌成本

        e_up = e.upper()
        if 'PT' in e_up or '铂' in e:
            mat_disp = 'PT950'
        elif 'K' in e_up:
            mat_disp = '18K'
        else:
            mat_disp = e or '18K'

        no += 1
        items.append({
            'no': no, 'row': r,
            '款号': c_str,
            '证书号': '',
            '条码号': '',
            '品名': str(b or '').strip(),
            '手寸': str(d or '').strip(),
            '材质颜色': mat_disp,
            '件数': int(f) if isinstance(f, (int, float)) else 1,
            '总重': g_wt,
            '主石重量_ct': main_w if main_w > 0 else None,
            '副石重量_合计': side_w_total,
            '镶嵌成本': math.ceil(ab_total),
            '_sheet': ws.title,
            '_date': date_str,
        })
    return items


# ==================== 聚水潭 (27 列) ====================
JST_COLS = [
    ('商品名称', 1), ('证书', 2), ('分类', 3), ('形状', 4),
    ('主石重量', 5), ('颜色等级', 6), ('净度', 7), ('切工', 8),
    ('抛光', 9), ('对称性', 10), ('荧光', 11), ('直径大小', 12),
    ('商品编码', 13), ('成本1', 14), ('成本2', 15), ('成本3', 16),
    ('成本价', 17), ('指圈号', 18), ('品名', 19), ('数量', 20),
    ('材质颜色', 21), ('总重', 22), ('副石重量', 23), ('商品简称', 24),
    ('品牌', 25), ('供应商名称', 26), ('供应商名', 27),
]

RED_FILL = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')


# v19.6: 品名细化 - 从工厂单原始品名提取品类关键词, 拼上分类
_PINMING_KEYWORDS = (
    '女戒', '男戒', '对戒', '排戒', '戒指',
    '手链', '手镯', '项链', '锁骨链', '颈链',
    '耳钉', '耳环', '耳夹', '耳线', '耳坠',
    '吊坠', '吊咀', '胸针', '脚链',
)


def _refine_pinming(category, raw_name):
    """category = '培育钻石' / '天然钻石' (来自 _classify_stone)
       raw_name = 工厂单品名 ('女戒' / '男戒' / '手链' / '手镯' / ...)
       输出: '培育钻石女戒' / '天然钻石手链' 等
       没识别到品类词的 → '{分类}戒指' (fallback)
    """
    if not raw_name:
        return f'{category}戒指'
    s = str(raw_name).strip()
    for kw in _PINMING_KEYWORDS:
        if kw in s:
            return f'{category}{kw}'
    return f'{category}{s}'


def _classify_stone(cert_no, attrs, item=None):
    """按证书类型/副石位置决定分类, 优先级从高到低:
       1. 工厂单证书号 LG 开头 → 培育钻石 (LG = Lab Grown)
       2. GIA 主石行证书列 = IGI → 培育钻石
       3. GIA 主石行证书列 = GIA (或其他非 IGI 值) → 天然钻石
       4. 没有主石行 (全散货 / 无匹配): 看工厂单副石列位
          - 副石2 金额 > 0 → 培育钻石 (布心副石2 = 培育散货, 单价 ~1120)
          - 副石1 金额 > 0 → 天然钻石 (布心副石1 = 天然散货, 单价 ~3200)
       5. 兜底 → 天然钻石 (真诚部门业务默认)
    """
    cert_no_up = str(cert_no or '').strip().upper()
    if cert_no_up.startswith('LG'):
        return '培育钻石'
    zheng = str((attrs or {}).get('证书') or '').strip().upper()
    if 'IGI' in zheng:
        return '培育钻石'
    if 'GIA' in zheng:
        return '天然钻石'
    # 没主石行 (可能只散货或飞书完全没匹配), 看工厂单副石1/副石2 谁有金额
    if item:
        s1 = item.get('_副石1金额') or 0
        s2 = item.get('_副石2金额') or 0
        if s2 > 0 and s2 >= s1:
            return '培育钻石'
        if s1 > 0:
            return '天然钻石'
    # 兜底
    return '天然钻石'


def build_jst_row(it, gia, factory_name):
    attrs = gia.get('attrs') or {}
    cost1 = gia.get('cost1') or 0
    cost2 = gia.get('cost2') or 0
    cost3 = it.get('镶嵌成本') or 0
    total = cost1 + cost2 + cost3
    # v20.3: category (培育钻石/天然钻石) 只用于品名前缀, "分类"字段统一"成品"
    category = _classify_stone(it.get('证书号'), attrs, it)
    return {
        '商品名称': it['款号'],
        '证书': attrs.get('证书'),
        '分类': '成品',   # v20.3: 统一"成品" (培育/天然区分在"品名"字段里)
        '形状': attrs.get('形状'),
        '主石重量': attrs.get('主石重量') or it.get('主石重量_ct'),
        '颜色等级': attrs.get('颜色等级'),
        '净度': attrs.get('净度'),
        '切工': attrs.get('切工'),
        '抛光': attrs.get('抛光'),
        '对称性': attrs.get('对称性'),
        '荧光': attrs.get('荧光'),
        '直径大小': None,
        '商品编码': _norm_code(it['证书号']) if it.get('证书号') else (it.get('条码号') or None),
        '成本1': round(cost1, 2) if cost1 > 0 else None,
        '成本2': round(cost2, 2) if cost2 > 0 else None,
        '成本3': round(cost3, 2) if cost3 > 0 else None,
        '成本价': round(total, 2) if total > 0 else None,
        '指圈号': it['手寸'],
        '品名': _refine_pinming(category, it.get('品名')),
        '数量': it['件数'],
        '材质颜色': it['材质颜色'],
        '总重': it['总重'],
        '副石重量': it['副石重量_合计'],
        '商品简称': it['款号'],
        '品牌': '真诚',
        '供应商名称': factory_name,
        '供应商名': factory_name,
        # v20.3 + v22.11: 红底警示 —— 三种情况任一满足就标红:
        #   1) 散货行 ≥ 2 (客户多次配石)
        #   2) 客户名下有额外主石 (回头客/对戒)
        #   3) v22.11: GIA 库存完全没找到这个客户 (无属性 + 无成本 + 无散货)
        '_红底': (gia.get('散货行数', 0) >= 2
                   or gia.get('主石额外行数', 0) >= 1
                   or (not attrs.get('证书')
                       and (gia.get('cost1') or 0) == 0
                       and (gia.get('cost2') or 0) == 0
                       and gia.get('散货行数', 0) == 0)),
    }


def gen_jst_xlsx(rows, out_path, sheet_title):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_title
    for col_name, col_idx in JST_COLS:
        ws.cell(row=1, column=col_idx).value = col_name
    for r_idx, row_data in enumerate(rows, start=2):
        for col_name, col_idx in JST_COLS:
            val = row_data.get(col_name)
            if val is not None:
                ws.cell(row=r_idx, column=col_idx).value = val
        if row_data.get('_红底'):
            for col_idx in (1, 24):
                ws.cell(row=r_idx, column=col_idx).fill = RED_FILL
    wb.save(out_path)


# ==================== 主 ====================
PARSERS = {'黛宝': parse_daibao, '布心': parse_buxin, '猛哥': parse_menggou, '二厂': parse_erchang}


def main():
    ap = argparse.ArgumentParser(
        description='天然钻客订聚水潭导出 (支持: 黛宝/布心/猛哥)',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog='示例: python3 natural.py 猛哥 郑总7月.xlsx --au 720 --pt 320 --out 猛哥_20260704.xlsx')
    ap.add_argument('factory', nargs='?', choices=list(PARSERS.keys()),
                    help='工厂名 (可省, 会按 sheet 名自动识别)')
    ap.add_argument('excel', help='工厂单 xlsx 路径')
    ap.add_argument('--au', type=float, required=True, help='当天足金金价 (元/克)')
    ap.add_argument('--pt', type=float, required=True, help='当天铂金金价 (元/克)')
    ap.add_argument('--out', help='聚水潭输出路径 (可选, 默认自动命名)')
    args = ap.parse_args()

    # 兼容 FEISHU_APP_ID/SECRET 和 FS_APP_ID/SECRET 两套变量名
    app_id = (os.environ.get('FEISHU_APP_ID') or os.environ.get('FS_APP_ID') or '').strip()
    app_secret = (os.environ.get('FEISHU_APP_SECRET') or os.environ.get('FS_APP_SECRET') or '').strip()
    if not (app_id and app_secret):
        print("请 export FEISHU_APP_ID / FEISHU_APP_SECRET (或 FS_APP_ID / FS_APP_SECRET)")
        sys.exit(1)

    xlsx_path = ensure_xlsx(args.excel)
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    factory = args.factory or detect_factory(wb)
    if not factory:
        print(f"❌ 无法识别工厂 (sheet 名: {wb.sheetnames}). 请把工厂名放在第一个参数, 例如: python3 natural.py 猛哥 ...")
        sys.exit(1)
    print(f"🏭 工厂: {factory}")

    print(f"\n=== Step 1: 解析工厂单 ===")
    items = PARSERS[factory](xlsx_path, args.au, args.pt)
    print(f"共 {len(items)} 件天然钻客订:")
    for it in items:
        print(f"  #{it['no']} 款号={it['款号']!r} 证书号={it['证书号']!r} "
              f"成色={it['材质颜色']} 镶嵌成本={it['镶嵌成本']}")

    if not items:
        print("⚠️  没天然钻客订件"); return

    print(f"\n=== Step 2: 查飞书 GIA 库存 ===")
    client = FeishuSheetClient(app_id, app_secret)
    all_sheets = client.list_sheets(GIA_TOKEN)
    order_sheets = [s for s in all_sheets if '订货' in s.get('title', '')]
    print(f"共 {len(order_sheets)} 个月度订货 sheet")

    for it in items:
        gia = search_gia(client, order_sheets, it['款号'], it['证书号'])
        it['_gia'] = gia
        attrs = gia['attrs']
        red = ' 🔴多散货' if gia['散货行数'] >= 2 else ''
        print(f"\n  #{it['no']} 款号={it['款号']!r}{red}")
        print(f"      成本1={gia['cost1']:.0f} 成本2={gia['cost2']:.0f} (散货{gia['散货行数']}条)")
        if attrs:
            print(f"      属性: {attrs.get('证书')} {attrs.get('形状')} {attrs.get('主石重量')}ct "
                  f"{attrs.get('颜色等级')} {attrs.get('净度')}")
        print(f"      命中: {gia['debug'][:200]}")

    print(f"\n=== Step 3: 生成聚水潭 ===")
    jst_rows = [build_jst_row(it, it['_gia'], factory) for it in items]
    ts = datetime.now().strftime('%Y%m%d_%H%M%S')
    date_str = datetime.now().strftime('%m-%d')
    base_dir = os.path.dirname(os.path.abspath(args.excel))
    if args.out:
        jst_out = args.out if os.path.isabs(args.out) else os.path.join(base_dir, args.out)
    else:
        jst_out = os.path.join(base_dir, f'聚水潭_{factory}{date_str}_{ts}.xlsx')
    gen_jst_xlsx(jst_rows, jst_out, sheet_title=f'{factory}{date_str}')
    print(f"✅ {jst_out}")
    red_count = sum(1 for r in jst_rows if r.get('_红底'))
    if red_count:
        print(f"⚠️  {red_count} 件多散货, 客户名已标红")
    print(f"\n🎉 完成!")


if __name__ == '__main__':
    main()
