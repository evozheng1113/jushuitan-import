"""4 家工厂解析 + 回写 (网站完整版 v23.11)
- parse_*: 解析工厂账单 → items
- write_*: 把客户名/利润/利率回写到工厂账单 → _完成.xlsx
"""
import openpyxl, math, re, os, subprocess, shutil, zipfile
from datetime import datetime, date


def _safe_copy(src, dst):
    shutil.copy(src, dst)
    os.chmod(dst, 0o644)


def _col_letter(n):
    s = ''
    while n > 0:
        n, r = divmod(n - 1, 26)
        s = chr(65 + r) + s
    return s


def _col_num(letters):
    n = 0
    for ch in letters:
        n = n * 26 + (ord(ch.upper()) - 64)
    return n


def _fmt_num_for_text(val):
    if isinstance(val, int):
        return str(val)
    if isinstance(val, float):
        if val == int(val):
            return str(int(val))
        return f"{val:.2f}".rstrip('0').rstrip('.')
    return str(val)


def _make_cell_xml(cell_ref, val, cached=None, preserve_attrs='', force_general=False):
    attrs = re.sub(r'\s*t="[^"]*"', '', preserve_attrs)
    if force_general:
        attrs = re.sub(r'\s*s="[^"]*"', '', attrs)
        if isinstance(val, (int, float)):
            text = _fmt_num_for_text(val)
            esc = text.replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')
            return f'<c r="{cell_ref}"{attrs} t="inlineStr"><is><t>{esc}</t></is></c>'
        elif isinstance(val, str) and not val.startswith('='):
            esc = val.replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')
            return f'<c r="{cell_ref}"{attrs} t="inlineStr"><is><t>{esc}</t></is></c>'
    if isinstance(val, str) and val.startswith('='):
        v_xml = f'<v>{cached}</v>' if cached is not None else ''
        return f'<c r="{cell_ref}"{attrs}><f>{val[1:]}</f>{v_xml}</c>'
    elif isinstance(val, (int, float)):
        return f'<c r="{cell_ref}"{attrs} t="n"><v>{val}</v></c>'
    else:
        esc = str(val).replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')
        return f'<c r="{cell_ref}"{attrs} t="inlineStr"><is><t>{esc}</t></is></c>'


_CELL_TAG_RE = re.compile(r'<c r="([A-Z]+)(\d+)"[^>]*?(?:/>|>.*?</c>)', re.DOTALL)


def _insert_cell_in_order(row_content, new_cell_xml, col_num):
    insert_pos = None
    for m in _CELL_TAG_RE.finditer(row_content):
        existing_col = _col_num(m.group(1))
        if existing_col > col_num:
            insert_pos = m.start()
            break
    if insert_pos is None:
        return row_content + new_cell_xml
    return row_content[:insert_pos] + new_cell_xml + row_content[insert_pos:]


def _reorder_row_cells(row_content):
    cells = []
    for m in _CELL_TAG_RE.finditer(row_content):
        cells.append((_col_num(m.group(1)), m.start(), m.end(), m.group(0)))
    if len(cells) < 2:
        return row_content
    already_sorted = all(cells[i][0] < cells[i+1][0] for i in range(len(cells)-1))
    if already_sorted:
        return row_content
    first_start = cells[0][1]
    last_end = cells[-1][2]
    sorted_cells = sorted(cells, key=lambda x: x[0])
    new_cells_str = ''.join(c[3] for c in sorted_cells)
    return row_content[:first_start] + new_cells_str + row_content[last_end:]


def _patch_sheet_xml(xml, mods):
    by_row = {}
    for (r, c), v in mods.items():
        by_row.setdefault(r, []).append((c, _col_letter(c), v))
    for row_num, changes in by_row.items():
        changes.sort(key=lambda x: x[0])
        row_pattern = re.compile(rf'(<row r="{row_num}"[^>]*>)(.*?)(</row>)', re.DOTALL)
        m = row_pattern.search(xml)
        if not m:
            continue
        row_start, row_content, row_end = m.group(1), m.group(2), m.group(3)
        new_content = row_content
        for col_n, col_letter_s, v in changes:
            opts = {}
            if isinstance(v, tuple):
                if len(v) >= 3:
                    val, cached, opts = v[0], v[1], v[2]
                elif len(v) == 2:
                    val, cached = v
                else:
                    val, cached = v[0], None
            else:
                val, cached = v, None
            cell_ref = f'{col_letter_s}{row_num}'
            cell_re = re.compile(rf'<c r="{cell_ref}"([^>]*?)(?:/>|>(.*?)</c>)')
            ex = cell_re.search(new_content)
            preserve = ex.group(1) if ex else ''
            new_cell = _make_cell_xml(cell_ref, val, cached, preserve,
                                      force_general=opts.get('force_general', False))
            if ex:
                new_content = cell_re.sub(new_cell, new_content, count=1)
            else:
                new_content = _insert_cell_in_order(new_content, new_cell, col_n)
        new_content = _reorder_row_cells(new_content)
        xml = xml.replace(m.group(0), row_start + new_content + row_end)
    return xml


def _resolve_sheet_xml_path(xlsx_path, sheet_title):
    with zipfile.ZipFile(xlsx_path) as z:
        wb_xml = z.read('xl/workbook.xml').decode('utf-8')
        rels_xml = z.read('xl/_rels/workbook.xml.rels').decode('utf-8')
    sheet_m = re.search(rf'<sheet[^>]*name="{re.escape(sheet_title)}"[^>]*r:id="(rId\d+)"', wb_xml)
    if not sheet_m:
        sheet_m = re.search(rf'<sheet[^>]*r:id="(rId\d+)"[^>]*name="{re.escape(sheet_title)}"', wb_xml)
    if not sheet_m:
        return 'xl/worksheets/sheet1.xml'
    rid = sheet_m.group(1)
    rel_m = re.search(rf'<Relationship[^>]*Id="{rid}"[^>]*Target="([^"]+)"', rels_xml)
    if not rel_m:
        rel_m = re.search(rf'<Relationship[^>]*Target="([^"]+)"[^>]*Id="{rid}"', rels_xml)
    if not rel_m:
        return 'xl/worksheets/sheet1.xml'
    target = rel_m.group(1)
    if target.startswith('/'):
        return target.lstrip('/')
    if target.startswith('worksheets/'):
        return 'xl/' + target
    return 'xl/' + target


def _patch_xlsx_cells(src_path, dst_path, modifications, sheet_name='sheet1', sheet_title=None):
    _safe_copy(src_path, dst_path)
    if sheet_title:
        target = _resolve_sheet_xml_path(dst_path, sheet_title)
    else:
        target = f'xl/worksheets/{sheet_name}.xml'
    tmp = dst_path + '.tmp'
    with zipfile.ZipFile(dst_path) as zin:
        with zipfile.ZipFile(tmp, 'w', zipfile.ZIP_DEFLATED) as zout:
            for n in zin.namelist():
                if n == target:
                    data = zin.read(n).decode('utf-8')
                    data = _patch_sheet_xml(data, modifications)
                    zout.writestr(n, data.encode('utf-8'))
                else:
                    zout.writestr(n, zin.read(n))
    os.replace(tmp, dst_path)


def _ensure_xlsx(path):
    if path.endswith('.xls') and not path.endswith('.xlsx'):
        out = path.rsplit('.', 1)[0] + '.xlsx'
        if not os.path.exists(out):
            try:
                subprocess.run(['soffice', '--headless', '--convert-to', 'xlsx',
                                path, '--outdir', os.path.dirname(out) or '.'],
                               check=True, timeout=30)
            except (subprocess.CalledProcessError, FileNotFoundError, subprocess.TimeoutExpired):
                raise RuntimeError("无法处理 .xls 文件,请先用 Excel/WPS 另存为 .xlsx 再上传")
        return out
    return path


def _fmt_profit_rate(v):
    if v is None:
        return None
    try:
        return f"{float(v) * 100:.1f}%"
    except (TypeError, ValueError):
        return str(v)


def _is_silver(material):
    if not material:
        return False
    s = str(material).upper()
    return 'S925' in s or '925' in s or '银' in str(material)


def _num(v):
    return v if isinstance(v, (int, float)) else 0


def _normalize_order(s):
    if not s:
        return ''
    return re.sub(r'-+', '-', str(s).strip())


def _unwrap_datetime_order(v):
    if isinstance(v, (datetime, date)):
        y = v.year
        if 2000 <= y < 2050:
            return f"{y - 2000}-{v.month}-{v.day}"
    return v


def _profit(val):
    return (val, None, {'force_general': True})


# ============ 工厂 A: 雅希 (广州厂) ============
def parse_A(excel_path, pt_price, au_price, sheet_name=None, **kw):
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb[wb.sheetnames[0]]
    items, current_rows = [], []
    for r in range(6, ws.max_row + 1):
        a = ws.cell(row=r, column=1).value
        if isinstance(a, (int, float)):
            if current_rows:
                items.append(_aggregate_A(ws, current_rows, pt_price, au_price))
            current_rows = [r]
        elif current_rows:
            current_rows.append(r)
    if current_rows:
        items.append(_aggregate_A(ws, current_rows, pt_price, au_price))

    processed_recycles = set()
    recycle_by_order = {}
    for it in items:
        if it['类别'] == '旧金回收' and it.get('下单编号'):
            key = str(it['下单编号']).strip()
            if key:
                recycle_by_order.setdefault(key, []).append(it)
    for it in items:
        if it['类别'] == '客户单' and it.get('下单编号'):
            key = str(it['下单编号']).strip()
            cands = [r for r in recycle_by_order.get(key, []) if id(r) not in processed_recycles]
            if cands:
                rec = cands[0]
                it['镶嵌成本'] = math.ceil((it['镶嵌成本'] or 0) + (rec['镶嵌成本'] or 0))
                processed_recycles.add(id(rec))

    merged = []
    i = 0
    while i < len(items):
        cur = items[i]
        nxt = items[i + 1] if i + 1 < len(items) else None
        if (cur['类别'] == '客户单' and nxt and nxt['类别'] == '旧金回收'
                and id(nxt) not in processed_recycles):
            cur['镶嵌成本'] = math.ceil((cur['镶嵌成本'] or 0) + (nxt['镶嵌成本'] or 0))
            processed_recycles.add(id(nxt))
            merged.append(cur)
            merged.append(nxt)
            i += 2
        else:
            merged.append(cur)
            i += 1
    return merged


def _aggregate_A(ws, rows, pt_price, au_price):
    r0 = rows[0]
    no = int(ws.cell(row=r0, column=1).value)
    order = ws.cell(row=r0, column=3).value
    invoice = ws.cell(row=r0, column=4).value
    cert = ws.cell(row=r0, column=5).value
    style = ws.cell(row=r0, column=6).value or ''
    material = ws.cell(row=r0, column=7).value
    qty = ws.cell(row=r0, column=8).value or 1
    total_weight = ws.cell(row=r0, column=9).value

    sum_M, sum_N = 0.0, 0.0
    for r in rows:
        sum_M += _num(ws.cell(row=r, column=13).value)
        sum_N += _num(ws.cell(row=r, column=14).value)

    main_w = _num(ws.cell(row=r0, column=18).value)
    side_w = _num(ws.cell(row=r0, column=23).value)
    if not side_w and len(rows) > 1:
        for rr in rows[1:]:
            side_w += _num(ws.cell(row=rr, column=23).value)

    invoice_str, order_str = str(invoice or ''), str(order or '')
    style_str = str(style or '')
    if '修理' in style_str:
        cat = '修理'
    elif '退石' in style_str or '退钱' in style_str:
        cat = '退还'
    elif '旧金回收' in str(cert or ''):
        cat = '旧金回收'
    elif order_str.startswith('2026-'):
        cat = '现货'
    elif re.search(r'[A-Z]-XH-', order_str.upper()):
        # v22.7: C 列含 A-XH- / B-XH- ... 即 XH 单号 (可能换行分隔多个)
        cat = '现货'
    elif invoice_str.rstrip().endswith('-XH'):
        # v22.7: D 列 -XH 结尾 (如 260718-XH), 说明这件是现货, E 列可能有多个证书号
        cat = '现货'
    elif '真诚' in invoice_str or '真诚' in order_str:
        cat = '部门-真诚'
    elif not order_str and not str(cert or ''):
        cat = '内部-跳过'
    else:
        cat = '客户单'

    is_silver = material == 'S925'

    if material == 'PT950':
        gp = pt_price
    elif material in ('18K', '14K', '10K', '24K'):
        gp = au_price
    elif is_silver:
        gp = ws.cell(row=r0, column=15).value or 23
    else:
        gp = 0

    cost = 0
    if cat == '旧金回收':
        ab = ws.cell(row=r0, column=28).value
        if isinstance(ab, (int, float)) and ab < 0:
            cost = math.ceil(ab)
        else:
            gold_cost = sum_M * (pt_price or 0) + sum_N * (au_price or 0)
            s = g = l = z = o = 0
            for r in rows:
                s += _num(ws.cell(row=r, column=21).value)
                g += _num(ws.cell(row=r, column=24).value)
                l += _num(ws.cell(row=r, column=25).value)
                z += _num(ws.cell(row=r, column=26).value)
                o += _num(ws.cell(row=r, column=27).value)
            cost = math.ceil(gold_cost + s + g + l + z + o)
    elif cat in ('修理', '内部-跳过', '退还'):
        cost = 0
    elif is_silver:
        ab = ws.cell(row=r0, column=28).value
        if isinstance(ab, (int, float)):
            cost = math.ceil(ab)
        else:
            p = _num(ws.cell(row=r0, column=16).value)
            s = g = l = z = o = 0
            for r in rows:
                s += _num(ws.cell(row=r, column=21).value)
                g += _num(ws.cell(row=r, column=24).value)
                l += _num(ws.cell(row=r, column=25).value)
                z += _num(ws.cell(row=r, column=26).value)
                o += _num(ws.cell(row=r, column=27).value)
            cost = math.ceil(p + s + g + l + z + o)
    else:
        gold_cost = sum_M * (pt_price or 0) + sum_N * (au_price or 0)
        s = g = l = z = o = 0
        for r in rows:
            s += _num(ws.cell(row=r, column=21).value)
            g += _num(ws.cell(row=r, column=24).value)
            l += _num(ws.cell(row=r, column=25).value)
            z += _num(ws.cell(row=r, column=26).value)
            o += _num(ws.cell(row=r, column=27).value)
        cost = math.ceil(gold_cost + s + g + l + z + o)

    fly_key = f'A-{order_str.strip()}' if cat == '客户单' and order else None
    return {'no': no, 'rows': rows, '类别': cat, '下单编号': order, '单号': invoice, '证书编号': cert,
            '品名': style, '成色': material, '件数': qty, '金价': gp, '镶嵌成本': cost,
            '总重': total_weight,
            '主石重量': main_w if main_w > 0 else None,
            '副石重量': side_w if side_w > 0 else None,
            '折铂金': sum_M, '折足金': sum_N, '_is_silver': is_silver,
            '_pt_price': pt_price, '_au_price': au_price,
            '飞书匹配键': fly_key, '飞书客户名': None}


def write_A(excel_path, items, out_path):
    modifications = {}
    for it in items:
        r = it['rows'][0]
        is_silver = it.get('_is_silver')
        sum_M = it.get('折铂金') or 0
        sum_N = it.get('折足金') or 0
        pt = it.get('_pt_price') or 0
        au = it.get('_au_price') or 0

        write_gold = it['类别'] in ('客户单', '现货', '部门-真诚') and not is_silver
        if write_gold and (sum_M > 0 or sum_N > 0):
            if sum_M > 0 and sum_N == 0:
                modifications[(r, 15)] = pt
                cached = round(pt * sum_M, 4)
                modifications[(r, 16)] = (f'=O{r}*M{r}', cached)
            elif sum_N > 0 and sum_M == 0:
                modifications[(r, 15)] = au
                cached = round(au * sum_N, 4)
                modifications[(r, 16)] = (f'=O{r}*N{r}', cached)
            else:
                modifications[(r, 15)] = pt
                cached = round(pt * sum_M + au * sum_N, 4)
                modifications[(r, 16)] = cached

        if it['类别'] in ('客户单', '现货', '部门-真诚') and it.get('镶嵌成本') and not is_silver:
            modifications[(r, 28)] = it['镶嵌成本']

        ac_map = {'现货': '现货', '修理': '修理', '旧金回收': '旧金回收',
                  '部门-真诚': '真诚', '内部-跳过': '内部', '退还': '退还'}
        ac = ac_map.get(it['类别'])
        if not ac and it['类别'] == '客户单':
            ac = it.get('飞书客户名') or '[待填]'
        if ac:
            modifications[(r, 29)] = ac

        # 现货 AD = ceil(AB / H)
        if it['类别'] == '现货' and it.get('镶嵌成本') and it.get('件数'):
            try:
                qty_n = int(it['件数']) if it['件数'] else 1
                if qty_n > 0:
                    unit_cost = math.ceil(float(it['镶嵌成本']) / qty_n)
                    modifications[(r, 30)] = unit_cost
            except (TypeError, ValueError, ZeroDivisionError):
                pass

        if it['类别'] == '客户单':
            profit = it.get('飞书利润')
            if isinstance(profit, (int, float)):
                modifications[(r, 30)] = _profit(profit)
            rate = _fmt_profit_rate(it.get('飞书利润率'))
            if rate is not None:
                modifications[(r, 31)] = _profit(rate)
    _patch_xlsx_cells(excel_path, out_path, modifications)


# ============ 工厂 B: 倾诚 ============
def parse_B(excel_path, sheet_name=None, **kw):
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb[wb.sheetnames[-1]]
    REPAIR = ['换扣', '旧扣抵扣', '加链', '换链', '换细链', '换粗链']
    items = []
    sheet_idx = wb.sheetnames.index(ws.title)
    for r in range(9, ws.max_row + 1):
        a = ws.cell(row=r, column=1).value
        if not isinstance(a, (int, float)):
            continue
        no = int(a)
        pinming = ws.cell(row=r, column=2).value
        invoice = ws.cell(row=r, column=3).value
        invoice = _unwrap_datetime_order(invoice)
        material = ws.cell(row=r, column=5).value
        qty = ws.cell(row=r, column=6).value
        total_weight = ws.cell(row=r, column=7).value
        total = ws.cell(row=r, column=28).value or 0
        # v20: 抓主石列 L (第 12 列) 作为证书编号
        #      "厂配"/"库石"/"客料"/中文名 → 排除, 只留 GIA/IGI 号 (纯字母数字, 长度 >=6)
        main_stone_raw = ws.cell(row=r, column=12).value
        cert_no = None
        if main_stone_raw is not None:
            ms = str(main_stone_raw).strip()
            # 无中文 + 长度 >=6 视作证书号
            if ms and len(ms) >= 6 and not any('一' <= ch <= '鿿' for ch in ms):
                cert_no = ms
        if not pinming and not invoice and not material and total == 0:
            continue
        inv = str(invoice or '').strip()
        if any(k in str(pinming or '') for k in REPAIR):
            cat = '维修'
        elif not inv:
            cat = '现货'
        elif re.match(r'^[A-Z]-XH-.+$', inv):
            # v20: {工厂code}-XH-* 是工厂下单出的现货 (A=雅希 B=二厂 D=黛宝 E=猛哥)
            cat = '现货'
        elif re.match(r'^\d+-\d+-\d+$', inv) or re.match(r'^B-\d+-\d+-\d+$', inv):
            cat = '客户单'
        elif inv.lower().startswith('tb'):
            cat = '客户单'
        elif re.match(r'^(\d+分|\d+号|[小中大])$', inv):
            cat = '现货'
        elif inv in ('星河款戒指', '粉蓝宝', '小帆鱼', '红绳', '鱼美人'):
            cat = '现货'
        else:
            cat = '客户单'
        fly_key = None
        if cat == '现货' and re.match(r'^[A-Z]-XH-.+$', inv):
            # 现货 XH 单号 = 匹配成品新单的键 (完整 inv)
            fly_key = inv
        elif cat == '客户单':
            if inv.startswith('B-'):
                fly_key = inv
            elif re.match(r'^\d+-\d+-\d+$', inv):
                fly_key = f'B-{inv}'
            else:
                fly_key = inv
        items.append({'no': no, 'rows': [r], '类别': cat, '品名': pinming, '单号': invoice,
                      '成色': material, '件数': qty, '镶嵌成本': math.ceil(total),
                      '总重': total_weight,
                      '证书编号': cert_no,   # v20: 主石 GIA/IGI 号 (兜底匹配用)
                      '飞书匹配键': fly_key,
                      '_sheet_idx': sheet_idx, '_sheet_title': ws.title})
    return items


def write_B(excel_path, items, out_path):
    modifications = {}
    sheet_title = items[0].get('_sheet_title') if items else None
    for it in items:
        r = it['rows'][0]
        ac = None
        if it['类别'] == '现货':
            ac = '现货'
        elif it['类别'] == '维修':
            ac = it.get('飞书客户名') or it.get('飞书匹配键') or '[待填]'
        elif it['类别'] == '客户单':
            ac = it.get('飞书客户名') or '[待填]'
        if ac:
            modifications[(r, 29)] = ac
        if it['类别'] == '客户单':
            profit = it.get('飞书利润')
            if isinstance(profit, (int, float)):
                modifications[(r, 30)] = _profit(profit)
            rate = _fmt_profit_rate(it.get('飞书利润率'))
            if rate is not None:
                modifications[(r, 31)] = _profit(rate)
    _patch_xlsx_cells(excel_path, out_path, modifications,
                      sheet_name='sheet1', sheet_title=sheet_title)


# ============ 工厂 D: 黛宝 ============
def _detect_D_columns(ws):
    """扫描黛宝出货单 r4 表头, 返回字段列号.
       v19.9: 新老表格差异 (新表 42 列, 老表 39 列, 每列左移 1-2 位)
       字段: 序号 条码号 款号 件数 手寸 名称 成色 总重(g) 净金重(g) 损耗 连耗重(g) 折足 补口价 补口费 主石 副石 总金额
    """
    COL = {}
    for c in range(1, ws.max_column + 1):
        v = str(ws.cell(row=4, column=c).value or '').replace('\n', '').replace(' ', '').strip()
        if not v: continue
        if v == '序号': COL['序号'] = c
        elif v == '条码号': COL['条码号'] = c
        elif v == '款号': COL['款号'] = c
        elif v == '件数': COL['件数'] = c
        elif v == '手寸': COL['手寸'] = c
        elif v == '名称': COL['名称'] = c
        elif v == '成色': COL['成色'] = c
        elif v.startswith('总重'): COL['总重'] = c
        elif v == '折足': COL['折足'] = c
        elif v == '主石': COL['主石'] = c
        elif v == '副石': COL['副石'] = c
        elif v == '总金额': COL['总金额'] = c
    return COL


def parse_D(excel_path, pt_price, au_price, **kw):
    excel_path = _ensure_xlsx(excel_path)
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb.active
    COL = _detect_D_columns(ws)
    # 关键字段兜底 (万一表头识别失败): 用老表位置作 fallback
    col_序 = COL.get('序号', 1)
    col_条码 = COL.get('条码号', 2)
    col_款 = COL.get('款号', 3)
    col_件 = COL.get('件数', 4)
    col_名 = COL.get('名称', 7)
    col_成 = COL.get('成色', 8)
    col_总重 = COL.get('总重', 10)
    col_折足 = COL.get('折足', 15)
    col_主石 = COL.get('主石', 19)
    col_总额 = COL.get('总金额', 38)

    items = []
    for r in range(6, ws.max_row + 1):
        a = ws.cell(row=r, column=col_序).value
        if not isinstance(a, (int, float)):
            continue
        no = int(a)
        barcode = ws.cell(row=r, column=col_条码).value
        kuanhao = ws.cell(row=r, column=col_款).value
        name = ws.cell(row=r, column=col_名).value
        material = ws.cell(row=r, column=col_成).value
        total_weight = ws.cell(row=r, column=col_总重).value
        zhezu = ws.cell(row=r, column=col_折足).value or 0
        AL = ws.cell(row=r, column=col_总额).value or 0
        # 证书号 = 主石列的 GIA 号 (旧代码抓 D=件数=1, 是错的)
        main_stone_no = ws.cell(row=r, column=col_主石).value
        cert = str(main_stone_no).strip() if main_stone_no else None

        # v20.8: 恢复原始简单规则 (用户昨天没问题的版本)
        #   - {工厂code}-XH-* → 现货 (工厂下单)
        #   - 数字-数字-数字 (如 7-6-2) → 客户单
        #   - "成品款式" 或空 → 现货 (工厂做的标品, 用证书号匹配成品新单)
        #   - 其他 (客户名/tb订单号等) → 客户单
        # v19.9/v20.6/v20.7 加的"客户名列判断"/"成品款式+主石=客户单"都是错的, 删掉
        ks = str(kuanhao or '').strip()
        has_main_stone = bool(main_stone_no)
        if re.match(r'^[A-Z]-XH-.+$', ks):
            cat, fly_key = '现货', ks
        elif re.match(r'^\d+-\d+-\d+$', ks):
            cat, fly_key = '客户单', f'D-{ks}'
        elif ks == '成品款式' or not ks:
            # 款号"成品款式" = 现货 (用证书号做匹配键, 兜底用条码号)
            cat, fly_key = '现货', cert
        else:
            cat, fly_key = '客户单', ks

        ms = str(material or '').upper()
        is_silver = _is_silver(material)
        if is_silver:
            gp = 0
            cost = math.ceil(AL)
        else:
            if 'PT950' in ms:
                gp = pt_price
            else:
                gp = au_price
            cost = math.ceil(zhezu * gp + AL) if gp else math.ceil(AL)
        items.append({'no': no, 'rows': [r, r + 1], '类别': cat, '条码号': barcode, '款号': kuanhao,
                      '证书编号': cert, '品名': name, '成色': material, '件数': 1, '金价': gp,
                      '折足': zhezu, '工石费': AL, '镶嵌成本': cost,
                      '总重': total_weight, '_is_silver': is_silver,
                      '_D_COL': COL,
                      '飞书匹配键': fly_key, '飞书客户名': None})
    return items


def write_D(excel_path, items, out_path):
    excel_path = _ensure_xlsx(excel_path)
    modifications = {}
    for it in items:
        r = it['rows'][0]
        gp = it['金价']
        COL = it.get('_D_COL') or {}
        # 总金额列 (回写公式的目标) — 新表可能是 39, 老表是 38
        am_col = COL.get('总金额', 38)
        # 折足列 (公式引用) — 新表 M(13), 老表 O(15)
        zz_col = COL.get('折足', 15)
        # 总金额列的**后 3 列** = 客户名/利润/利润率 (跟 write_D 原逻辑保持)
        name_col = am_col + 1
        profit_col = am_col + 2
        rate_col = am_col + 3
        zz_letter = openpyxl.utils.get_column_letter(zz_col)
        am_letter = openpyxl.utils.get_column_letter(am_col)

        if it.get('_is_silver') or not gp:
            if it.get('镶嵌成本'):
                modifications[(r, am_col)] = it['镶嵌成本']
        else:
            cached = round(gp * (it['折足'] or 0) + (it['工石费'] or 0), 4)
            modifications[(r, am_col)] = (f'={gp}*{zz_letter}{r}+{am_letter}{r}', cached)
        an = '现货' if it['类别'] == '现货' else (it.get('飞书客户名') or '[待填]')
        modifications[(r, name_col)] = an
        if it['类别'] == '客户单':
            profit = it.get('飞书利润')
            if isinstance(profit, (int, float)):
                modifications[(r, profit_col)] = _profit(profit)
            rate = _fmt_profit_rate(it.get('飞书利润率'))
            if rate is not None:
                modifications[(r, rate_col)] = _profit(rate)
    _patch_xlsx_cells(excel_path, out_path, modifications)


# ============ 工厂 E: 猛哥 ============
ORDER_LIKE = re.compile(r'^\d+-+\d+-+\d+$')
BATCH_LIKE = re.compile(r'^([A-Za-z]+\d+)[-—]?$')


def _detect_E_columns(ws):
    keys = ['序号', '单号', '编码', '品名', '成色', '手寸', '件数',
            '总重量', '净金重', '损耗', '含耗金重',
            '足金金重', '折足金', '金价', '金值',
            '配件重', '配件费', '主石', '副石',
            '镶主石费', '版蜡', '工艺费', '工费', '总金额']
    found = {}
    for c in range(1, 40):
        v = ws.cell(row=4, column=c).value
        if not v: continue
        s = str(v).strip()
        for key in keys:
            if key in s and key not in found:
                found[key] = c
                break
    if '折足金' not in found and '足金金重' in found:
        found['折足金'] = found['足金金重']
    if '品名' in found and ('成色' not in found or found['成色'] == found['品名']):
        next_col = found['品名'] + 1
        used = set(v for k, v in found.items() if k != '成色')
        if next_col not in used:
            found['成色'] = next_col
    if '件数' not in found and '成色' in found:
        found['件数'] = found['成色'] + 2
    if '总重量' not in found and '件数' in found:
        found['总重量'] = found['件数'] + 1
    if '主石' in found:
        found['主石重量'] = found['主石'] + 1
        # v15.5: 主石金额位置 — 若有副石表头, 用副石-1 (最后一列); 否则 hardcode +4
        if '副石' in found and found['副石'] > found['主石']:
            found['主石金额'] = found['副石'] - 1
        else:
            found['主石金额'] = found['主石'] + 4
    if '副石' in found:
        found['副石重量'] = found['副石'] + 1
        # 副石金额位置: 有镶主石费的话, 副石金额 = 镶主石费 - 1
        if '镶主石费' in found and found['镶主石费'] > found['副石']:
            found['副石金额'] = found['镶主石费'] - 2   # 副石金额 / 镶石费 / 镶主石费
            found['镶石费'] = found['镶主石费'] - 1
        else:
            found['副石金额'] = found['副石'] + 4
            found['镶石费'] = found['副石'] + 5
    if '工艺费' in found:
        found['版蜡'] = found['工艺费'] - 1
    defaults = {
        '序号': 1, '单号': 2, '品名': 3, '成色': 4, '件数': 6, '总重量': 7,
        '含耗金重': 10, '折足金': 11, '金价': 12, '金值': 13,
        '配件费': 15, '主石重量': 18, '主石金额': 20,
        '副石重量': 23, '副石金额': 26, '镶石费': 27,
        '镶主石费': 28, '版蜡': 29, '工艺费': 30, '工费': 31, '总金额': 32,
    }
    for k, v in defaults.items():
        if k not in found:
            found[k] = v
    return found


def parse_E(excel_path, pt_price, au_price, sheet_name=None, default_material=None, **kw):
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb[wb.sheetnames[-1]]
    sheet_idx = wb.sheetnames.index(ws.title)
    items = []
    auto_no_counter = 0
    COL = _detect_E_columns(ws)

    for r in range(6, ws.max_row + 1):
        a = ws.cell(row=r, column=COL['序号']).value
        b = ws.cell(row=r, column=COL['单号']).value
        # v15.4: 天然钻石模板下单单号写在"编码"列, "单号"列空 → 用编码列兜底
        if not b and COL.get('编码') and COL['编码'] != COL['单号']:
            code_val = ws.cell(row=r, column=COL['编码']).value
            if code_val:
                b = code_val
        b_str = str(b or '').strip()
        c_cell = ws.cell(row=r, column=COL['品名']).value
        d_cell = ws.cell(row=r, column=COL['成色']).value

        if not d_cell and default_material:
            d_cell = default_material

        no = None
        if isinstance(a, (int, float)):
            no = int(a)
        elif isinstance(a, str) and a.strip():
            astr = a.strip()
            stripped = astr.rstrip('-').rstrip('—')
            try:
                no = int(stripped)
            except ValueError:
                m = BATCH_LIKE.match(astr)
                if m:
                    no = m.group(1)
                elif stripped == '19楼':
                    no = '19楼'

        if no is None:
            if b_str and ORDER_LIKE.match(b_str):
                auto_no_counter += 1
                no = f"自动-{auto_no_counter}"
            elif c_cell and d_cell:
                auto_no_counter += 1
                no = f"现货-{auto_no_counter}"
            else:
                continue

        invoice = b
        pinming = c_cell
        material = d_cell
        qty = ws.cell(row=r, column=COL['件数']).value or 1
        total_weight = ws.cell(row=r, column=COL['总重量']).value
        han_hao = _num(ws.cell(row=r, column=COL['含耗金重']).value)
        factory_K = ws.cell(row=r, column=COL['折足金']).value
        factory_L = ws.cell(row=r, column=COL['金价']).value
        factory_M = ws.cell(row=r, column=COL['金值']).value
        peijian_fei  = _num(ws.cell(row=r, column=COL['配件费']).value)
        zhushi_jin   = _num(ws.cell(row=r, column=COL['主石金额']).value)
        fushi_jin    = _num(ws.cell(row=r, column=COL['副石金额']).value)
        xiangshi_fei = _num(ws.cell(row=r, column=COL['镶石费']).value)
        xzhushi_fei  = _num(ws.cell(row=r, column=COL['镶主石费']).value)
        banla        = _num(ws.cell(row=r, column=COL['版蜡']).value)
        gongyi       = _num(ws.cell(row=r, column=COL['工艺费']).value)
        gongfei      = _num(ws.cell(row=r, column=COL['工费']).value)
        main_w       = _num(ws.cell(row=r, column=COL['主石重量']).value)
        side_w       = _num(ws.cell(row=r, column=COL['副石重量']).value)

        if not invoice and not material:
            continue

        is_silver = _is_silver(material)
        mat_s = str(material or '').lower()

        if is_silver:
            ratio = 0
        elif 'pt' in mat_s:
            ratio = 0.952
        else:
            ratio = 0.755

        if isinstance(factory_K, (int, float)) and factory_K != 0:
            zhe_zu = factory_K
        elif han_hao and ratio:
            zhe_zu = round(han_hao * ratio, 5)
        else:
            zhe_zu = 0

        if isinstance(factory_L, (int, float)) and factory_L > 0:
            gp = factory_L
        elif 'pt' in mat_s:
            gp = pt_price or 0
        else:
            gp = au_price or 0

        if isinstance(factory_M, (int, float)) and factory_M > 0:
            jin_zhi = factory_M
        elif zhe_zu and gp:
            jin_zhi = round(zhe_zu * gp, 5)
        else:
            jin_zhi = 0

        if is_silver:
            factory_AF = _num(ws.cell(row=r, column=COL['总金额']).value)
            cost = math.ceil(factory_AF) if factory_AF else math.ceil(
                peijian_fei + zhushi_jin + fushi_jin + xiangshi_fei +
                xzhushi_fei + banla + gongyi + gongfei
            )
        else:
            cost = math.ceil(jin_zhi + peijian_fei + zhushi_jin + fushi_jin +
                              xiangshi_fei + xzhushi_fei + banla + gongyi + gongfei)

        invoice_s = str(invoice or '').strip()
        normalized = _normalize_order(invoice_s)

        if no == '19楼' or invoice_s == '19楼':
            cat, fly_key, customer = '部门-真诚', None, '真诚'
        elif re.match(r'^[A-Z]-XH-.+$', invoice_s):
            # v20: {工厂code}-XH-* 是工厂下单的现货 (E-XH-*)
            cat, fly_key, customer = '现货', invoice_s, '现货'
        elif not invoice_s:
            cat, fly_key, customer = '现货', None, '现货'
        elif ORDER_LIKE.match(invoice_s):
            cat, fly_key, customer = '客户单', f'E-{normalized}', None
        else:
            cat, fly_key, customer = '客户单', invoice_s, None

        items.append({'no': no, 'rows': [r], '类别': cat, '单号': invoice, '品名': pinming,
                      '成色': material, '件数': qty,
                      '总重': total_weight,
                      '主石重量': main_w if main_w > 0 else None,
                      '副石重量': side_w if side_w > 0 else None,
                      '含耗金重': han_hao, '折足金': zhe_zu, '金价': gp, '金值': jin_zhi,
                      '_factory_K_filled': isinstance(factory_K, (int, float)) and factory_K != 0,
                      '_factory_L_filled': isinstance(factory_L, (int, float)) and factory_L > 0,
                      '_factory_M_filled': isinstance(factory_M, (int, float)) and factory_M > 0,
                      '镶嵌成本': cost, '_is_silver': is_silver,
                      '_E_COL': COL,
                      '飞书匹配键': fly_key, '飞书客户名': customer,
                      '_sheet_idx': sheet_idx, '_sheet_title': ws.title})

    # v15.5 + v22.1: 合并连续同单号+同品名的行 (一对耳钉工厂占两行, 业务上算 1 件)
    # v22.1 修正: 猛哥工厂出货是一只一只的成本, 需相加合并成一对总成本
    #        (旧版只保留第一行, 导致成本少一半)
    merged = []
    for it in items:
        if (merged and it['类别'] == '客户单' and merged[-1]['类别'] == '客户单'
                and it.get('飞书匹配键') and it.get('飞书匹配键') == merged[-1].get('飞书匹配键')
                and it.get('品名') == merged[-1].get('品名')):
            # 合并到前一个 item: rows 追加 + 成本相加
            merged[-1]['rows'].extend(it['rows'])
            merged[-1]['镶嵌成本'] = (merged[-1].get('镶嵌成本') or 0) + (it.get('镶嵌成本') or 0)
        else:
            merged.append(it)
    return merged


def write_E(excel_path, items, out_path):
    if not items:
        return
    COL = items[0].get('_E_COL') or {}
    modifications = {}
    sheet_title = items[0].get('_sheet_title')
    af_col = COL.get('总金额', 32)
    name_col = af_col + 1
    profit_col = af_col + 2
    rate_col = af_col + 3
    for it in items:
        is_silver = it.get('_is_silver')
        # v15.5: 遍历所有 rows (合并后的多行件, 每行都要填)
        for r in it['rows']:
            if not is_silver and not it.get('_factory_K_filled') and it.get('折足金'):
                modifications[(r, COL.get('折足金', 11))] = it['折足金']
            if not is_silver and not it.get('_factory_L_filled') and it.get('金价'):
                modifications[(r, COL.get('金价', 12))] = it['金价']
            if not is_silver and not it.get('_factory_M_filled') and it.get('金值'):
                k_letter = _col_letter(COL.get('折足金', 11))
                l_letter = _col_letter(COL.get('金价', 12))
                modifications[(r, COL.get('金值', 13))] = (f'={k_letter}{r}*{l_letter}{r}', it['金值'])
            if it.get('镶嵌成本'):
                modifications[(r, af_col)] = it['镶嵌成本']

            customer = it.get('飞书客户名')
            if not customer:
                if it['类别'] == '客户单':
                    customer = '[待填]'
                elif it['类别'] == '现货':
                    customer = '现货'
                elif it['类别'] == '部门-真诚':
                    customer = '真诚'
            if customer:
                modifications[(r, name_col)] = customer

            if it['类别'] == '客户单':
                profit = it.get('飞书利润')
                if isinstance(profit, (int, float)):
                    modifications[(r, profit_col)] = _profit(profit)
                rate = _fmt_profit_rate(it.get('飞书利润率'))
                if rate is not None:
                    modifications[(r, rate_col)] = _profit(rate)
    _patch_xlsx_cells(excel_path, out_path, modifications,
                      sheet_name='sheet1', sheet_title=sheet_title)


PARSERS = {'A': parse_A, 'B': parse_B, 'D': parse_D, 'E': parse_E}
WRITERS = {'A': write_A, 'B': write_B, 'D': write_D, 'E': write_E}


def get_parser(code):
    return PARSERS[code]


def get_writer(code):
    return WRITERS[code]
