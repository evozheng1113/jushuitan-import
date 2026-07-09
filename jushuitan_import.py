"""聚水潭批量入库模板生成 v9
v9: 1) 工厂账单主石重/副石重优先
    2) 商品名称简化成品类 (戒指/耳环/项链/吊坠...)
    3) 主石类别: 天然钻石单子 → "天然钻石"
    4) 指圈号只对戒指带出
    5) 天然钻石单子: 商品名称前加 "天然"
"""
import openpyxl
import os
import re


COL = {
    '成色': 1, '主石类别': 2, '品名': 3, '形状': 4, '总重': 5,
    '主石重量': 6, '副石重量': 7, '颜色等级': 8, '净度': 9, '荧光': 10,
    '商品编码': 11, '商品名称': 12, '指圈号': 13, '数量': 14,
    '折后价': 15, '直径大小': 16, '成本1': 17, '成本2': 18, '成本价': 19,
    '分类': 20, '品牌': 21, '供应商名称': 22, '供应商名': 23,
    '客户名称': 24,   # v15.1 新增
}

SUPPLIER_MAP = {
    'A': '雅希-广州厂',
    'B': 'JC-二厂',
    'D': '黛宝',
    'E': '创艺-猛哥',
}


_CATEGORY_KEYWORDS = [
    ('钥匙吊', '吊坠'),
    ('锁骨链', '项链'),
    ('对戒', '戒指'),
    ('耳钉', '耳环'),
    ('耳坠', '耳环'),
    ('耳夹', '耳环'),
    ('耳线', '耳环'),
    ('手环', '手镯'),
    ('吊咀', '吊坠'),
    ('吊坠', '吊坠'),
    ('戒指', '戒指'),
    ('耳环', '耳环'),
    ('项链', '项链'),
    ('手链', '手链'),
    ('手镯', '手镯'),
    ('胸针', '胸针'),
    ('脚链', '脚链'),
    ('颈链', '项链'),
]


def extract_category(name):
    if name is None:
        return None
    s = str(name)
    for keyword, category in _CATEGORY_KEYWORDS:
        if keyword in s:
            return category
    return None


def _normalize_dashes(s):
    return re.sub(r'-+', '-', str(s))


def normalize_cert_to_code(cert, factory_code=None, is_spot=False):
    """
    v22.2: 现货 (is_spot=True) → -2 后缀, 客户单 → -1 后缀
    """
    if cert is None:
        return None
    s = str(cert).strip()
    if not s:
        return None
    if '/' in s:
        return s

    s = _normalize_dashes(s)
    s = re.sub(r'\s+[a-zA-Z]+(?:-\d+)?$', '', s).rstrip()

    suffix = '-2' if is_spot else '-1'

    dash_count = s.count('-')
    if dash_count >= 2:
        return s
    if dash_count == 1:
        prefix, _, last = s.rpartition('-')
        if last in ('1', '2'):
            # 已带 -1/-2, 按 is_spot 重置成正确后缀
            s = prefix
        else:
            return s
    if factory_code == 'D' and not s.startswith('D'):
        return f'D{s}{suffix}'
    return f'{s}{suffix}'


def normalize_material(material):
    if material is None:
        return None
    s = str(material).strip()
    if not s:
        return None
    up = s.upper()
    if 'PT' in up:
        return 'PT950'
    if 'S925' in up or '925' in up or '银' in s:
        return 'S925'
    repl = {'k白': '18K白', 'K白': '18K白',
            'k黄': '18K黄', 'K黄': '18K黄',
            'k红': '18K红', 'K红': '18K红',
            'k玫': '18K玫', 'K玫': '18K玫'}
    for old, new in repl.items():
        if old in s and '18K' not in s.upper():
            return s.replace(old, new)
    if 'K' in up and '18' not in s:
        return '18' + s.upper().replace('K', 'K')
    return s


def parse_main_stone(stone_text):
    if not stone_text:
        return (None, None, None)
    s = str(stone_text).strip()
    weight = None
    m = re.search(r'(\d+(?:\.\d+)?)\s*分', s)
    if m:
        weight = round(float(m.group(1)) / 100, 4)
    else:
        m = re.search(r'^(\d+(?:\.\d+)?)\s*(?:ct)?', s)
        if m:
            try:
                w = float(m.group(1))
                if 0.01 <= w <= 10:
                    weight = w
            except ValueError:
                pass
    color = None
    m = re.search(r'\b([D-K])\b', s)
    if m:
        color = m.group(1)
    clarity = None
    m = re.search(r'\b(VVS[12]|VS[12]|SI[12]|IF|FL)\b', s)
    if m:
        clarity = m.group(1)
    return (weight, color, clarity)


def build_row_from_item(item, factory_code, feishu_cert,
                        feishu_luozuan=0, feishu_peishi=0,
                        feishu_main_stone=None, feishu_ring_size=None,
                        total_weight=None, is_natural=False):
    # v22.2: 现货 -2, 客户单 -1
    is_spot = item.get('类别') == '现货'
    code = normalize_cert_to_code(feishu_cert, factory_code=factory_code, is_spot=is_spot)
    if not code:
        order = item.get('下单编号') or item.get('单号')
        if order and factory_code != 'D':
            norm_order = _normalize_dashes(str(order).strip())
            code = f'{factory_code}-{norm_order}'
        else:
            code = None

    row = {c: None for c in COL}
    row['主石类别'] = '天然钻石' if is_natural else '培育钻石'
    row['数量'] = 1
    row['分类'] = '成品1'
    row['品牌'] = '倾诚'
    row['供应商名称'] = SUPPLIER_MAP.get(factory_code, '')
    row['供应商名'] = SUPPLIER_MAP.get(factory_code, '')
    row['商品编码'] = code
    row['成色'] = normalize_material(item.get('成色'))

    pinming = item.get('品名')
    category = extract_category(pinming)

    if factory_code == 'D':
        row['品名'] = '戒指'
        if category is None:
            category = '戒指'
    else:
        row['品名'] = category or pinming

    if category:
        row['商品名称'] = ('天然' if is_natural else '') + category
    else:
        if pinming:
            row['商品名称'] = ('天然' if is_natural else '') + str(pinming)

    if total_weight is not None and total_weight > 0:
        row['总重'] = round(float(total_weight), 4)
    elif item.get('总重'):
        row['总重'] = round(float(item['总重']), 4)

    factory_main_w = item.get('主石重量') or item.get('主石重')
    factory_side_w = item.get('副石重量') or item.get('副石重') or item.get('配石重量') or item.get('配石重')
    if isinstance(factory_main_w, (int, float)) and factory_main_w > 0:
        row['主石重量'] = round(float(factory_main_w), 4)
    if isinstance(factory_side_w, (int, float)) and factory_side_w > 0:
        row['副石重量'] = round(float(factory_side_w), 4)

    if feishu_main_stone:
        w, color, clarity = parse_main_stone(feishu_main_stone)
        if w is not None and not row.get('主石重量'):
            row['主石重量'] = w
        if color:
            row['颜色等级'] = color
        if clarity:
            row['净度'] = clarity

    if feishu_ring_size and category == '戒指':
        row['指圈号'] = str(feishu_ring_size).strip()

    row['成本2'] = item.get('镶嵌成本')
    cost1 = (feishu_luozuan or 0) + (feishu_peishi or 0)
    row['成本1'] = round(cost1, 2) if cost1 > 0 else None
    row['成本价'] = round((row['成本1'] or 0) + (row['成本2'] or 0), 2) if (row['成本1'] or row['成本2']) else None

    # v15.1: 客户名称 (同终端 process.py 逻辑)
    # - 客户单: 飞书客户名 (含已退款"X已退款做现货" 这种情况) / 没查到则 [待填]
    # - 现货:   "现货"
    # - 真诚:   "真诚"
    cat = item.get('类别')
    customer = item.get('飞书客户名')
    if cat == '现货':
        customer = customer or '现货'
    elif cat == '部门-真诚':
        customer = customer or '真诚'
    elif cat == '客户单':
        customer = customer or '[待填]'
    if customer:
        row['客户名称'] = customer

    return row


def generate_or_append(output_path, item_rows, template_path=None):
    added = 0
    replaced = 0

    if os.path.exists(output_path):
        wb = openpyxl.load_workbook(output_path)
        ws = wb.active
        existing = {}
        for r in range(2, ws.max_row + 1):
            code = ws.cell(row=r, column=COL['商品编码']).value
            if code:
                existing[str(code).strip()] = r
        next_row = ws.max_row + 1
    elif template_path and os.path.exists(template_path):
        wb = openpyxl.load_workbook(template_path)
        ws = wb.active
        if ws.max_row > 1:
            ws.delete_rows(2, ws.max_row - 1)
        existing = {}
        next_row = 2
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        for col_name, col_idx in COL.items():
            ws.cell(row=1, column=col_idx).value = col_name
        existing = {}
        next_row = 2

    for item in item_rows:
        code = item.get('商品编码')
        if not code:
            continue
        code_key = str(code).strip()
        if code_key in existing:
            r = existing[code_key]
            replaced += 1
        else:
            r = next_row
            next_row += 1
            existing[code_key] = r
            added += 1
        for col_name, value in item.items():
            if col_name in COL and value is not None:
                ws.cell(row=r, column=COL[col_name]).value = value

    wb.save(output_path)
    return (added, replaced)
